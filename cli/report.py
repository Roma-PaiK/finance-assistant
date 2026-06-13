"""
Block 7+8 — Monthly Spend Report + Budget Tracking

Usage:
  python report.py --month YYYY-MM              # terminal table for one month
  python report.py --month YYYY-MM --budget     # overlay budget targets + alerts
  python report.py --month YYYY-MM --csv        # + export CSV
  python report.py --month YYYY-MM --excel      # + export Excel
  python report.py --year YYYY                  # full-year view
  python report.py --month YYYY-MM --compare YYYY-MM  # explicit MoM compare
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import argparse
import yaml
from datetime import datetime, date
from collections import defaultdict

from core.db import query, init_db
from core.analytics import spend_by_category, top_merchants, savings_rate

BUDGET_PATH = os.path.join(os.path.dirname(__file__), "..", "config", "budget.yaml")


# ── helpers ──────────────────────────────────────────────────────────────────

def _prev_month(ym: str) -> str:
    """'2025-03' → '2025-02', '2025-01' → '2024-12'"""
    y, m = int(ym[:4]), int(ym[5:7])
    if m == 1:
        return f"{y-1:04d}-12"
    return f"{y:04d}-{m-1:02d}"


def _fmt_inr(amount: float) -> str:
    """Format float as ₹ with Indian comma grouping."""
    neg = amount < 0
    amount = abs(amount)
    s = f"{amount:,.0f}"
    # Indian grouping: last 3 then groups of 2
    parts = s.split(",")
    if len(parts) <= 1:
        result = s
    else:
        # re-do grouping properly
        int_part = str(int(abs(amount)))
        if len(int_part) <= 3:
            result = int_part
        else:
            result = int_part[-3:]
            int_part = int_part[:-3]
            while int_part:
                result = int_part[-2:] + "," + result
                int_part = int_part[:-2]
    return ("−" if neg else "") + "₹" + result


def _delta_str(curr: float, prev: float) -> str:
    if prev == 0:
        return "  new"
    diff = curr - prev
    pct = (diff / prev) * 100
    sign = "+" if diff >= 0 else "−"
    return f"{sign}₹{abs(diff):,.0f} ({sign}{abs(pct):.0f}%)"


def _bar(amount: float, max_amount: float, width: int = 20) -> str:
    if max_amount == 0:
        return ""
    filled = int((amount / max_amount) * width)
    return "█" * filled + "░" * (width - filled)


# ── Budget helpers ────────────────────────────────────────────────────────────

def _load_budget() -> dict:
    if not os.path.exists(BUDGET_PATH):
        return {}
    with open(BUDGET_PATH) as f:
        return yaml.safe_load(f) or {}


def _budget_status(actual: float, target: float, alert_pct: float, warn_pct: float) -> str:
    """Return status string: OVER / WARN / ok"""
    if actual > target * (alert_pct / 100):
        return "OVER"
    if actual > target * (warn_pct / 100):
        return "WARN"
    return "ok"


def _print_budget_report(month: str, cats: dict[str, float]):
    cfg = _load_budget()
    if not cfg:
        print("  No budget.yaml found — skipping budget overlay.")
        return

    targets: dict[str, float] = cfg.get("monthly_targets", {})
    fixed: list[str] = cfg.get("fixed_categories", [])
    alert_pct = cfg.get("thresholds", {}).get("alert_pct", 110)
    warn_pct  = cfg.get("thresholds", {}).get("warn_pct", 90)

    STATUS_ICON = {"OVER": "🔴", "WARN": "🟡", "ok": "🟢", "fixed": "──", "no_target": "  "}

    col_w = 28
    print(f"\n{'─'*72}")
    print(f"  Budget vs Actual — {month}")
    print(f"  {'─'*col_w}  {'Actual':>10}  {'Budget':>10}  {'Used':>7}  Status")
    print(f"  {'─'*col_w}  {'─'*10}  {'─'*10}  {'─'*7}  ──────")

    alerts = []
    for cat, actual in sorted(cats.items(), key=lambda x: -x[1]):
        if cat in fixed:
            icon = STATUS_ICON["fixed"]
            target_str = "  fixed"
            used_str = ""
            status_str = "(fixed)"
        elif cat in targets:
            target = targets[cat]
            pct = (actual / target * 100) if target else 0
            status = _budget_status(actual, target, alert_pct, warn_pct)
            icon = STATUS_ICON[status]
            target_str = _fmt_inr(target)
            used_str = f"{pct:.0f}%"
            status_str = status
            if status in ("OVER", "WARN"):
                alerts.append((icon, cat, actual, target, pct))
        else:
            icon = STATUS_ICON["no_target"]
            target_str = "  —"
            used_str = ""
            status_str = "(no target)"

        print(f"  {icon} {cat:<{col_w-2}}  {_fmt_inr(actual):>10}  {target_str:>10}  {used_str:>7}  {status_str}")

    # Categories with budget targets but zero spend this month
    for cat, target in sorted(targets.items()):
        if cat not in cats:
            print(f"  🟢 {cat:<{col_w-2}}  {'₹0':>10}  {_fmt_inr(target):>10}  {'0%':>7}  ok")

    # Alert summary
    if alerts:
        print(f"\n  ⚠  Budget alerts:")
        for icon, cat, actual, target, pct in alerts:
            over = actual - target
            print(f"     {icon} {cat}: {_fmt_inr(actual)} spent vs {_fmt_inr(target)} budget "
                  f"({pct:.0f}% — {_fmt_inr(over)} over)")

    print(f"{'─'*72}")


# ── DB queries ────────────────────────────────────────────────────────────────

def _get_other_count(month: str) -> int:
    """Count of 'Other' category rows still in review queue."""
    rows = query("""
        SELECT COUNT(*) as n
        FROM transactions
        WHERE month = ?
          AND category = 'Other'
          AND transaction_type = 'genuine_spend'
    """, (month,))
    return rows[0]["n"] if rows else 0


def _get_month_summary(month: str) -> dict:
    """Total inflow, outflow, genuine spend for a month."""
    rows = query("""
        SELECT
            txn_type,
            transaction_type,
            SUM(amount) as total
        FROM transactions
        WHERE month = ?
        GROUP BY txn_type, transaction_type
    """, (month,))
    result = {"inflow": 0.0, "genuine_spend": 0.0, "internal_out": 0.0, "cc_settlement": 0.0}
    for r in rows:
        if r["txn_type"] == "credit":
            result["inflow"] += r["total"]
        elif r["txn_type"] == "debit":
            if r["transaction_type"] == "genuine_spend":
                result["genuine_spend"] += r["total"]
            elif r["transaction_type"] in ("cc_settlement", "internal_transfer"):
                result["internal_out"] += r["total"]
    return result


def _get_year_monthly(year: str) -> list[dict]:
    """Per-month genuine spend totals for a full year."""
    rows = query("""
        SELECT month, SUM(amount) as total
        FROM transactions
        WHERE month LIKE ?
          AND transaction_type = 'genuine_spend'
          AND txn_type = 'debit'
        GROUP BY month
        ORDER BY month
    """, (f"{year}-%",))
    return [dict(r) for r in rows]


def _print_savings_section(data: dict, month: str):
    """Print Block 10 savings awareness section."""
    year = month[:4]
    print(f"\n{'─'*64}")
    print(f"  Savings & Investments")
    print(f"{'─'*64}")

    if data["salary"]:
        print(f"  Salary inflow     : {_fmt_inr(data['salary'])}")
    else:
        print(f"  Salary inflow     : — (no SBI salary credit found this month)")

    if data["sip_month"]:
        print(f"  SIP outflow       : {_fmt_inr(data['sip_month'])}  ({data['sip_count']} deductions)")
    else:
        print(f"  SIP outflow       : — (no acc_bob_sip deductions this month)")

    print(f"  SIP YTD ({year})    : {_fmt_inr(data['sip_ytd'])}")
    print(f"  Genuine spend     : {_fmt_inr(data['genuine_spend'])}")

    if data["savings_rate"] is not None:
        rate = data["savings_rate"]
        note = "⚠ low" if rate < 20 else ("✓ good" if rate >= 40 else "")
        print(f"  Savings rate      : {rate:.1f}%  {note}")
        print(f"  ⚠  Indicative only — excludes tax, insurance, inter-account flows")
    else:
        print(f"  Savings rate      : — (salary not found; load SBI statement first)")

    print(f"{'─'*64}")


def _get_year_category_breakdown(year: str) -> dict[str, dict[str, float]]:
    """category → {month → total} for a full year."""
    rows = query("""
        SELECT month, category, SUM(amount) as total
        FROM transactions
        WHERE month LIKE ?
          AND transaction_type = 'genuine_spend'
          AND txn_type = 'debit'
        GROUP BY month, category
        ORDER BY month, total DESC
    """, (f"{year}-%",))
    result: dict[str, dict[str, float]] = defaultdict(dict)
    for r in rows:
        result[r["category"]][r["month"]] = r["total"]
    return dict(result)


# ── formatters ────────────────────────────────────────────────────────────────

def _print_month_report(month: str, compare_month: str | None = None, show_budget: bool = False):
    cats = spend_by_category(month)
    prev_cats = spend_by_category(compare_month or _prev_month(month))
    merchants = top_merchants(month, n=5)
    other_count = _get_other_count(month)
    summary = _get_month_summary(month)
    compare_label = compare_month or _prev_month(month)

    # Header
    print()
    print(f"{'─'*64}")
    print(f"  Monthly Spend Report — {month}")
    print(f"{'─'*64}")

    # Summary line
    print(f"  Genuine spend : {_fmt_inr(summary['genuine_spend'])}")
    print(f"  Inflow        : {_fmt_inr(summary['inflow'])}")
    if summary["cc_settlement"] or summary["internal_out"]:
        print(f"  Internal/CC   : {_fmt_inr(summary['internal_out'])}  (excluded from spend)")
    print(f"{'─'*64}")

    # Category table
    max_total = max(cats.values()) if cats else 1
    col_w = 28
    print(f"  {'Category':<{col_w}}  {'Amount':>10}  {'vs ' + compare_label:>20}  Bar")
    print(f"  {'─'*col_w}  {'─'*10}  {'─'*20}  {'─'*20}")

    for cat, total in sorted(cats.items(), key=lambda x: -x[1]):
        prev = prev_cats.get(cat, 0)
        delta = _delta_str(total, prev)
        bar = _bar(total, max_total)
        print(f"  {cat:<{col_w}}  {_fmt_inr(total):>10}  {delta:>20}  {bar}")

    # Other row
    if other_count:
        print(f"\n  ⚠  {other_count} transaction(s) still categorised as 'Other' — run review.py")

    # Top merchants
    print(f"\n{'─'*64}")
    print(f"  Top merchants — {month}")
    print(f"  {'─'*col_w}  {'Amount':>10}  {'Txns':>6}")
    for m in merchants:
        name = (m["merchant"] or "Unknown")[:col_w]
        print(f"  {name:<{col_w}}  {_fmt_inr(m['total']):>10}  {m['txn_count']:>6}")

    print(f"{'─'*64}")

    # Budget overlay (Block 8)
    if show_budget:
        _print_budget_report(month, cats)

    # Savings awareness (Block 10) — always shown
    savings = savings_rate(month)
    _print_savings_section(savings, month)

    print()


def _print_year_report(year: str):
    monthly = _get_year_monthly(year)
    cat_breakdown = _get_year_category_breakdown(year)
    months_present = [r["month"] for r in monthly]

    print()
    print(f"{'─'*72}")
    print(f"  Year Report — {year}")
    print(f"{'─'*72}")

    # Monthly totals row
    print(f"\n  Monthly genuine spend:")
    for r in monthly:
        print(f"    {r['month']}   {_fmt_inr(r['total']):>12}")

    total_year = sum(r["total"] for r in monthly)
    avg = total_year / len(monthly) if monthly else 0
    print(f"    {'─'*20}")
    print(f"    Total  {_fmt_inr(total_year):>12}")
    print(f"    Avg/mo {_fmt_inr(avg):>12}")

    # Category breakdown by month
    if months_present:
        print(f"\n  Category breakdown:")
        short_months = [m[5:] for m in months_present]   # MM only
        header = f"  {'Category':<28}" + "".join(f"  {m:>8}" for m in short_months) + f"  {'Total':>10}"
        print(header)
        print(f"  {'─'*28}" + "  ────────" * len(months_present) + "  ──────────")

        all_cats = sorted(
            cat_breakdown.keys(),
            key=lambda c: -sum(cat_breakdown[c].values())
        )
        for cat in all_cats:
            month_totals = cat_breakdown[cat]
            row = f"  {cat:<28}"
            cat_total = 0.0
            for m in months_present:
                v = month_totals.get(m, 0.0)
                cat_total += v
                row += f"  {_fmt_inr(v):>8}" if v else "          "
            row += f"  {_fmt_inr(cat_total):>10}"
            print(row)

    print(f"{'─'*72}")
    print()


# ── CSV / Excel export ────────────────────────────────────────────────────────

def _export_month_csv(month: str, path: str):
    import csv
    cats = spend_by_category(month)
    prev_cats = spend_by_category(_prev_month(month))
    rows = []
    for cat, total in sorted(cats.items(), key=lambda x: -x[1]):
        prev = prev_cats.get(cat, 0)
        diff = total - prev
        pct = (diff / prev * 100) if prev else None
        rows.append({
            "month": month,
            "category": cat,
            "amount": round(total, 2),
            "prev_month_amount": round(prev, 2),
            "change_abs": round(diff, 2),
            "change_pct": round(pct, 1) if pct is not None else "",
        })
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    print(f"  CSV → {path}")


def _export_month_excel(month: str, path: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  openpyxl not installed — run: uv add openpyxl")
        return

    cats = spend_by_category(month)
    prev_cats = spend_by_category(_prev_month(month))
    top_m = top_merchants(month, n=10)
    summary = _get_month_summary(month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Spend {month}"

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")

    # Summary section
    ws.append(["Monthly Spend Report", month])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Genuine Spend", round(summary["genuine_spend"], 2)])
    ws.append(["Inflow", round(summary["inflow"], 2)])
    ws.append(["Internal / CC Settlement", round(summary["internal_out"], 2)])
    ws.append([])

    # Category table
    headers = ["Category", "Amount (₹)", f"Prev Month (₹)", "Change (₹)", "Change (%)"]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.fill = hdr_fill
        cell.font = hdr_font

    for cat, total in sorted(cats.items(), key=lambda x: -x[1]):
        prev = prev_cats.get(cat, 0)
        diff = total - prev
        pct = (diff / prev * 100) if prev else ""
        ws.append([cat, round(total, 2), round(prev, 2), round(diff, 2),
                   round(pct, 1) if pct != "" else ""])

    ws.append([])

    # Top merchants
    ws.append(["Top Merchants", "Amount (₹)", "Transactions"])
    for cell in ws[ws.max_row]:
        cell.fill = hdr_fill
        cell.font = hdr_font
    for m in top_m:
        ws.append([m["merchant"], round(m["total"], 2), m["txn_count"]])

    # Column widths
    ws.column_dimensions["A"].width = 32
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 16

    wb.save(path)
    print(f"  Excel → {path}")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Finance monthly spend report")
    parser.add_argument("--month", help="YYYY-MM to report on")
    parser.add_argument("--year", help="YYYY for full-year view")
    parser.add_argument("--compare", help="YYYY-MM to compare against (default: previous month)")
    parser.add_argument("--budget", action="store_true", help="Overlay budget targets + alerts")
    parser.add_argument("--csv", action="store_true", help="Export CSV")
    parser.add_argument("--excel", action="store_true", help="Export Excel (.xlsx)")
    parser.add_argument("--out", help="Output file path (optional, auto-named if not set)")
    args = parser.parse_args()

    if not args.month and not args.year:
        # Default: current month
        args.month = date.today().strftime("%Y-%m")

    init_db()

    if args.year and not args.month:
        _print_year_report(args.year)
        return

    if args.month:
        _print_month_report(args.month, compare_month=args.compare, show_budget=args.budget)

        if args.csv:
            out = args.out or f"report_{args.month}.csv"
            _export_month_csv(args.month, out)

        if args.excel:
            out = args.out or f"report_{args.month}.xlsx"
            _export_month_excel(args.month, out)


if __name__ == "__main__":
    main()
