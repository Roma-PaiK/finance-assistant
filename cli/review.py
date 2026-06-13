"""
Block 4 — Dry-Run Review Interface.

Two modes:

  EXPORT — pull transactions from DB → reviewable CSV/Excel
  ─────────────────────────────────────────────────────────
  uv run python review.py export [--month 2025-01] [--source acc_canara_daily]
                                  [--all] [--excel] [--out path]

  Flags:
    --month   filter by month (YYYY-MM); repeatable
    --source  filter by source_id; repeatable
    --all     include internal transfers (excluded by default)
    --excel   write .xlsx with category dropdown + conditional formatting
    --out     output path (default: review_YYYYMMDD_HHMMSS.csv / .xlsx)

  Sorting: Other category first → ascending confidence → date
  Columns: corrected_category (blank) + transfer_type (blank) prepended for editing.

  APPLY — read corrected CSV/Excel → UPDATE DB + corrections cache
  ─────────────────────────────────────────────────────────────────
  uv run python review.py apply <corrected_file> [--save]

  Matches rows by (date, amount, txn_type, source_id).
  Only rows with non-blank corrected_category are processed.
  transfer_type merge: same rules as import_corrections.py.
  --save commits changes; default is preview only.
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import yaml
import pandas as pd
from collections import Counter
from datetime import datetime as dt
from core.corrections_merge import merge_transfer_type as _merge_transfer_type

CONFIG_DIR      = os.path.join(os.path.dirname(__file__), "..", "config")
CATEGORIES_YAML = os.path.join(CONFIG_DIR, "categories.yaml")
ACCOUNTS_YAML   = os.path.join(CONFIG_DIR, "accounts.yaml")


# ── Shared helpers ────────────────────────────────────────────────────────────

def _load_categories() -> list[str]:
    with open(CATEGORIES_YAML) as f:
        cfg = yaml.safe_load(f)
    return list(cfg.get("categories", {}).keys())



# ── EXPORT ────────────────────────────────────────────────────────────────────

def cmd_export(args: list[str]):
    from core.db import query

    # Parse flags
    months   = _collect_flag(args, "--month")
    sources  = _collect_flag(args, "--source")
    include_internal = "--all" in args
    as_excel = "--excel" in args
    out_path = _flag_value(args, "--out")

    # Build query
    conditions = []
    params     = []
    if months:
        placeholders = ",".join("?" * len(months))
        conditions.append(f"month IN ({placeholders})")
        params.extend(months)
    if sources:
        placeholders = ",".join("?" * len(sources))
        conditions.append(f"source_id IN ({placeholders})")
        params.extend(sources)
    if not include_internal:
        conditions.append("is_internal_transfer = 0")

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    sql   = f"SELECT * FROM transactions {where} ORDER BY date"
    rows  = query(sql, tuple(params))

    if not rows:
        print("No rows matched. Check --month / --source values.")
        return

    df = pd.DataFrame(rows)

    # Add blank review columns
    df["corrected_category"] = ""
    df["transfer_type"]      = ""

    # Sort: Other first → low confidence → date
    df["_sort_other"] = (df["category"] == "Other").astype(int)
    df = df.sort_values(["_sort_other", "confidence", "date"], ascending=[False, True, True])
    df = df.drop(columns=["_sort_other"])

    # Column order — review cols first
    front = ["id", "date", "month", "source_id", "source_label",
             "corrected_category", "transfer_type",
             "category", "category_source", "confidence",
             "txn_type", "amount", "description", "raw_description",
             "canonical_merchant", "is_internal_transfer", "notes"]
    df = df[[c for c in front if c in df.columns]]

    # Default output name
    if not out_path:
        ts  = dt.now().strftime("%Y%m%d_%H%M%S")
        ext = ".xlsx" if as_excel else ".csv"
        tag = "_".join(months or sources or ["all"])
        out_path = f"review_{tag}_{ts}{ext}"

    if as_excel:
        _write_excel(df, out_path)
    else:
        df.to_csv(out_path, index=False)
        print(f"Exported {len(df)} rows → {out_path}")

    _print_export_summary(df)


def _write_excel(df: pd.DataFrame, out_path: str):
    """Write Excel with category dropdown + colour-coded confidence."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.worksheet.datavalidation import DataValidation

    categories = _load_categories()

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Review")
        wb = writer.book
        ws = writer.sheets["Review"]

        # Find corrected_category column index (1-based)
        col_letters = {cell.value: cell.column_letter for cell in ws[1]}
        cc_col = col_letters.get("corrected_category")
        cat_col = col_letters.get("category")
        conf_col = col_letters.get("confidence")

        if cc_col:
            # Data validation dropdown for corrected_category
            cat_formula = '"' + ",".join(categories) + '"'
            dv = DataValidation(type="list", formula1=cat_formula, allow_blank=True)
            dv.error      = "Use category from dropdown"
            dv.errorTitle = "Invalid category"
            dv.prompt     = "Select corrected category (blank = accept)"
            ws.add_data_validation(dv)
            dv.add(f"{cc_col}2:{cc_col}{ws.max_row}")

        # Colour rows: red = Other, amber = confidence < 0.6, green = high conf
        red_fill   = PatternFill("solid", fgColor="FFCCCC")
        amber_fill = PatternFill("solid", fgColor="FFE5B4")
        green_fill = PatternFill("solid", fgColor="CCFFCC")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cat_cell  = None
            conf_cell = None
            for cell in row:
                if ws.cell(1, cell.column).value == "category":
                    cat_cell = cell
                if ws.cell(1, cell.column).value == "confidence":
                    conf_cell = cell

            if cat_cell is None:
                continue
            category   = str(cat_cell.value or "")
            confidence = float(conf_cell.value or 0) if conf_cell else 1.0

            if category == "Other":
                fill = red_fill
            elif confidence < 0.6:
                fill = amber_fill
            elif confidence >= 0.9:
                fill = green_fill
            else:
                fill = None

            if fill:
                for cell in row:
                    cell.fill = fill

        # Freeze header + auto-width
        ws.freeze_panes = "A2"
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    print(f"Exported {len(df)} rows → {out_path}  (Excel w/ dropdown)")


def _print_export_summary(df: pd.DataFrame):
    other_count    = (df["category"] == "Other").sum()
    low_conf_count = (df["confidence"] < 0.6).sum() if "confidence" in df.columns else 0
    print(f"\n  {len(df)} rows | {other_count} Other | {low_conf_count} low-confidence (<0.6)")
    if "category" in df.columns:
        cats = Counter(df["category"])
        print("  Category breakdown:")
        for cat, n in sorted(cats.items(), key=lambda x: -x[1]):
            print(f"    {cat:<44} {n:>4}")


# ── APPLY ─────────────────────────────────────────────────────────────────────

def cmd_apply(args: list[str]):
    from core.db import get_connection
    from core.corrections_db import upsert, stats, init_corrections_table

    save  = "--save" in args
    files = [a for a in args if not a.startswith("--")]

    if not files:
        print("Usage: uv run python review.py apply <corrected_file> [--save]")
        sys.exit(1)

    valid_categories = set(_load_categories())
    init_corrections_table()

    all_changes = []

    for path in files:
        if not os.path.exists(path):
            print(f"Not found: {path}")
            continue

        if path.endswith(".xlsx") or path.endswith(".xls"):
            df = pd.read_excel(path, dtype=str).fillna("")
        else:
            df = pd.read_csv(path, dtype=str).fillna("")

        if "corrected_category" not in df.columns:
            print(f"No 'corrected_category' column in {path} — skipping")
            continue

        # Only rows where correction was given
        mask = df["corrected_category"].str.strip() != ""
        corrections = df[mask].copy()

        if corrections.empty:
            print(f"{path}: no corrections filled.")
            continue

        for _, row in corrections.iterrows():
            corrected_cat = row["corrected_category"].strip()
            transfer_type = (row.get("transfer_type", "") or "").strip()

            if corrected_cat not in valid_categories:
                print(f"  Unknown category '{corrected_cat}' — skipping row {row.get('date','?')}")
                continue

            final_cat, needs_review = _merge_transfer_type(corrected_cat, transfer_type)
            is_internal = final_cat.startswith("Internal Transfer")

            # Match key: prefer id column, fall back to composite
            txn_id = _safe_int(row.get("id", ""))
            merchant = (row.get("canonical_merchant", "") or "").strip()
            original_cat = (row.get("category", "") or "").strip()

            all_changes.append({
                "txn_id":        txn_id,
                "date":          row.get("date", ""),
                "amount":        row.get("amount", ""),
                "txn_type":      (row.get("txn_type", "") or "").strip(),
                "source_id":     (row.get("source_id", "") or "").strip(),
                "canonical_merchant": merchant,
                "original_cat":  original_cat,
                "final_cat":     final_cat,
                "is_internal":   is_internal,
                "needs_review":  needs_review,
                "description":   row.get("description", ""),
            })

    if not all_changes:
        print("No valid corrections found.")
        return

    _print_apply_summary(all_changes)

    if not save:
        print("\nPreview only. Run with --save to commit to DB + corrections cache.")
        return

    # Commit changes
    conn = get_connection()
    updated = 0
    cached  = 0

    for c in all_changes:
        if c["txn_id"]:
            conn.execute(
                "UPDATE transactions SET category = ?, is_internal_transfer = ? WHERE id = ?",
                (c["final_cat"], int(c["is_internal"]), c["txn_id"])
            )
            updated += 1
        else:
            # Fallback: match by composite key
            result = conn.execute(
                """UPDATE transactions SET category = ?, is_internal_transfer = ?
                   WHERE date = ? AND amount = ? AND txn_type = ? AND source_id = ?""",
                (c["final_cat"], int(c["is_internal"]),
                 c["date"], c["amount"], c["txn_type"], c["source_id"])
            )
            updated += result.rowcount

        if c["canonical_merchant"]:
            upsert(c["canonical_merchant"], c["final_cat"])
            cached += 1

    conn.commit()
    conn.close()

    s = stats()
    print(f"\nSaved.")
    print(f"  DB: {updated} rows updated")
    print(f"  Cache: {cached} merchant mappings upserted")
    print(f"  Cache total: {s['total_merchants']} merchants | {s['high_confidence_3plus']} high-confidence (3+)")

    flagged = [c for c in all_changes if c["needs_review"]]
    if flagged:
        print(f"\n  {len(flagged)} rows still flagged (transfer_type=unknown) — re-export and resolve:")
        for c in flagged:
            print(f"    {c['date']}  {c['description'][:40]}  ₹{c['amount']}")


def _print_apply_summary(changes: list[dict]):
    print(f"\n{'='*64}")
    print(f"  {len(changes)} corrections to apply")
    print(f"{'='*64}")
    for c in changes:
        arrow = f"{c['original_cat']} → {c['final_cat']}"
        label = c["canonical_merchant"] or c["description"][:30]
        flag  = "  ⚠️  needs_review" if c["needs_review"] else ""
        print(f"  {c['date']}  {label:<30}  {arrow}{flag}")
    print(f"{'='*64}")


# ── CLI arg helpers ───────────────────────────────────────────────────────────

def _collect_flag(args: list[str], flag: str) -> list[str]:
    """Collect all values after `flag` (repeatable)."""
    result = []
    i = 0
    while i < len(args):
        if args[i] == flag and i + 1 < len(args):
            result.append(args[i + 1])
            i += 2
        else:
            i += 1
    return result


def _flag_value(args: list[str], flag: str) -> str | None:
    """Return single value after `flag`, or None."""
    vals = _collect_flag(args, flag)
    return vals[0] if vals else None


def _safe_int(val) -> int | None:
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]
    if not args:
        print(__doc__)
        sys.exit(0)

    cmd  = args[0]
    rest = args[1:]

    if cmd == "export":
        cmd_export(rest)
    elif cmd == "apply":
        cmd_apply(rest)
    else:
        print(f"Unknown command: {cmd}")
        print("Usage: uv run python review.py export|apply ...")
        sys.exit(1)


if __name__ == "__main__":
    main()
